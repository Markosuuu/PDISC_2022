from tkinter import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os

from modules.LeerExcel import *

ventana = Tk()
ventana.title("Modulo python-excel")
ventana.geometry("300x300")
ventana.config(bg="#2F454D")
ventana.resizable(width=False, height=False)

# Ventana que agrega datos a los archivos Excel
def nueva():
    ventana_nueva = Toplevel(ventana)
    
    # Funciones
    def salir():
        ventana_nueva.destroy()
    
    def agregarDatos():
        nombre = nom.get().capitalize()
        apellido = ape.get().capitalize()
        dnii = dni.get()
        archi = archivo.get()

        datos = [nombre, apellido, dnii]

        # Si no existe el archivo lo crea y luego sigue para leerlo.
        if not os.path.exists(f'{archi}.xlsx'):
            wb = Workbook()
            ws = wb.active

            # Color de los bordes
            black = "000000"
            thin = Side(border_style="thin", color=black)

            # Color del cabezal del excel
            green = "8db600"
            for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
                for cell in rows:
                    cell.fill = PatternFill(
                        start_color=green, end_color=green, fill_type="solid")

            # Estilo de las primeras celdas
            ws['A1'] = "Nombre"
            ws['A1'].border = Border(top=thin, right=thin, bottom=thin, left=thin)
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

            ws['B1'] = "Apellido"
            ws['B1'].border = Border(top=thin, right=thin, bottom=thin, left=thin)
            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

            ws['C1'] = "DNI"
            ws['C1'].border = Border(top=thin, right=thin, bottom=thin, left=thin)
            ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
            

            wb.save(f'{archi}.xlsx')

        wb = load_workbook(f'{archi}.xlsx')

        ws = wb.active
        ws.append(datos)

        wb.save(f'{archi}.xlsx')

    # Ventana tkinter
    ventana_nueva.title("Formulario")
    ventana_nueva.geometry("500x250")
    ventana_nueva.config(bg="#41c73a")
    ventana_nueva.focus_set()

    # Variables
    archivo = StringVar()
    nom = StringVar()
    ape = StringVar()
    dni = StringVar()
    
    Label(ventana_nueva, 
        text="Cargar datos en excel", 
        bg="#186914", 
        fg="#ffffff", 
        width=500, 
        height=2
        ).pack()

    Label(ventana_nueva, 
        text="Ingrese el nombre del archivo: ", 
        bg="#41c73a", 
        font=("Arial", 12)
        ).place(relx=0.1, rely=0.2)
    Entry(ventana_nueva, textvariable=archivo).place(relx=0.6, rely=0.2)

    Label(ventana_nueva, 
        text="Ingrese su nombre: ", 
        bg="#41c73a", 
        font=("Arial", 12)
        ).place(relx=0.1, rely=0.35)
    Entry(ventana_nueva, textvariable=nom).place(relx=0.6, rely=0.35)

    Label(ventana_nueva, 
        text="Ingrese su apellido: ", 
        bg="#41c73a", 
        font=("Arial", 12)
        ).place(relx=0.1, rely=0.50)
    Entry(ventana_nueva, textvariable=ape).place(relx=0.6, rely=0.50)

    Label(ventana_nueva, 
        text="Ingrese su DNI: ", 
        bg="#41c73a", 
        font=("Arial", 12)
        ).place(relx=0.1, rely=0.65)
    Entry(ventana_nueva, textvariable=dni).place(relx=0.6, rely=0.65)

    Button(ventana_nueva, 
        text="Enviar datos", 
        command=agregarDatos,
        bg="#186914",
        fg="#ffffff",
        padx=10,
        pady=10
        ).place(relx=0.25, rely=0.8)

    Button(ventana_nueva, 
        text="Cerrar ventana", 
        command=salir,
        bg="#186914",
        fg="#ffffff",
        padx=10,
        pady=10
        ).place(relx=0.50, rely=0.8)

Button(ventana, 
    text="Añadir informacion al excel", 
    command=nueva, 
    height=2, 
    width=30, 
    bg="#011E2A", 
    fg="white"
    ).place(relx=0.13, rely=0.25)

Button(ventana, 
    text="Ver achivos excel", 
    command=leerExcel, 
    height=2, 
    width=18, 
    bg="#011E2A", 
    fg="white"
    ).place(relx=0.25, rely=0.50)

ventana.mainloop()
